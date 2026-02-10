from flask import Flask, render_template, request, jsonify
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook

# ==================================================
# APP & PATH SETUP (OS-SAFE)
# ==================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, "templates"),
    static_folder=os.path.join(BASE_DIR, "static")
)

# ❗ FIX: Windows path removed – now works on Render/Linux
DATA_PATH = os.path.join(BASE_DIR, "data")
FILE_PATH = os.path.join(DATA_PATH, "daily_sales.xlsx")
ITEM_FILE = os.path.join(BASE_DIR, "items_master.xlsx")

os.makedirs(DATA_PATH, exist_ok=True)

# ==================================================
# SAFE TNo GENERATOR
# ==================================================
def get_next_tno():
    if not os.path.exists(FILE_PATH):
        return 1

    try:
        df = pd.read_excel(FILE_PATH)
    except Exception:
        return 1

    if df.empty:
        return 1

    df.columns = [str(c).strip() for c in df.columns]

    if "TNo" not in df.columns:
        return 1

    last_tno = pd.to_numeric(df["TNo"], errors="coerce").max()
    if pd.isna(last_tno):
        return 1

    return int(last_tno) + 1

# ==================================================
# ROUTES
# ==================================================
@app.route("/")
def index():
    template_path = os.path.join(app.template_folder, "index.html")

    if not os.path.exists(template_path):
        return (
            f"ERROR: index.html not found.<br>"
            f"Expected path: {template_path}<br>"
            f"Files found: {os.listdir(app.template_folder) if os.path.exists(app.template_folder) else 'templates folder missing'}",
            500
        )

    return render_template(
        "index.html",
        tno=get_next_tno(),
        today=datetime.today().strftime("%d/%m/%Y")
    )


@app.route("/items")
def items():
    if not os.path.exists(ITEM_FILE):
        return jsonify([])
    df = pd.read_excel(ITEM_FILE)
    return jsonify(df.iloc[:, 0].dropna().astype(str).unique().tolist())

@app.route("/save", methods=["POST"])
def save():
    data = request.json or {}

    # ===============================
    # BACKEND VALIDATION
    # ===============================
    mandatory = ["date", "shop", "item", "total_kg"]
    for m in mandatory:
        if not data.get(m):
            return jsonify({
                "status": "error",
                "message": f"{m} is mandatory"
            }), 400

    # ===============================
    # FORCE TNo
    # ===============================
    data["TNo"] = get_next_tno()

    # ===============================
    # NUMERIC SANITIZATION
    # ===============================
    for k, v in data.items():
        try:
            data[k] = round(float(v), 2)
        except Exception:
            pass

    df_new = pd.DataFrame([data])

    if os.path.exists(FILE_PATH):
        df_old = pd.read_excel(FILE_PATH)
        df = pd.concat([df_old, df_new], ignore_index=True)
    else:
        df = df_new

    df.to_excel(FILE_PATH, index=False)

    # ===============================
    # EXCEL FORMATTING
    # ===============================
    wb = load_workbook(FILE_PATH)
    ws = wb.active
    ws.freeze_panes = "A2"

    for col in ws.iter_cols(min_row=2):
        for cell in col:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    wb.save(FILE_PATH)

    return jsonify({
        "status": "success",
        "next_tno": get_next_tno()
    })

@app.route("/records")
def records():
    if not os.path.exists(FILE_PATH):
        return jsonify([])
    df = pd.read_excel(FILE_PATH)
    return jsonify(df.to_dict(orient="records"))

@app.route("/record/<int:tno>")
def record(tno):
    if not os.path.exists(FILE_PATH):
        return jsonify({}), 404
    df = pd.read_excel(FILE_PATH)
    rec = df[df["TNo"] == tno]
    if rec.empty:
        return jsonify({}), 404
    return jsonify(rec.iloc[0].to_dict())

@app.route("/delete/<int:tno>", methods=["POST"])
def delete(tno):
    if not os.path.exists(FILE_PATH):
        return jsonify({"status": "not_found"}), 404
    df = pd.read_excel(FILE_PATH)
    df = df[df["TNo"] != tno]
    df.to_excel(FILE_PATH, index=False)
    return jsonify({"status": "deleted"})

# ==================================================
# MAIN
# ==================================================
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)
